# -*-coding:utf-8-*-
# Time:2017/9/21 19:02
# Author:YangYangJun


from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

import os
import time


def writeExcel():
    wb = Workbook()
    ws = wb.active
    count = 1
    with open("xx.txt","r",encoding="utf-8") as f:
        for row in f:
            w_row = [i.strip() for i in row.split("===>")]
            # print(w_row)
            ws.append(w_row)
            wb.save(filename="demo.xlsx")
            print(f"{count}ok")
            count+=1



def readExcel(ExcelFullName):
    wb = load_workbook(ExcelFullName)
    # wb = load_workbook(filename=ExcelFullName)

    # 获取当前活跃的worksheet,默认就是第一个worksheet
    # ws = wb.active
    # 当然也可以使用下面的方法
    # 获取所有表格(worksheet)的名字
    sheets = wb.get_sheet_names()
    print(sheets)
    # # 第一个表格的名称
    sheet_first = sheets[0]
    # # 获取特定的worksheet
    #
    ws = wb.get_sheet_by_name(sheet_first)
    print("***")
    print(sheet_first)
    print(ws.title)
    print("^^^")
    # 获取表格所有行和列，两者都是可迭代的

    rows = ws.rows
    print(rows)

    columns = ws.columns

    # 迭代所有的行

    for row in rows:
        line = [col.value for col in row]

        print(line)

    # 通过坐标读取值

    print(ws['A1'].value)  # A表示列,1表示行

    print(ws.cell(row=1, column=1).value)




if __name__ == '__main__':
    writeExcel()

